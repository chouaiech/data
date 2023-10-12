import datetime
import os
import hashlib
import logging
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Border, Side, PatternFill
from util.app_configuration import AppConfiguration
from util.identifier import Identifier


class XlsxFileConnector:

    def __init__(self):
        self._log = logging.getLogger("XlsxFileConnector")
        self._config = AppConfiguration()
        self._fileprefix = "edp_statistics_"
        self._filesuffix = ".xlsx"
        self._filepath = self._config.resource_path

    def is_file_existing(self, filename):
        return os.path.isfile(self._filepath + filename)

    def create_file(self, remote_ip):
        filename = self._create_filename(remote_ip)
        wb = Workbook()
        wb.save(self._filepath + filename)
        return filename

    def delete_all_xlsx_files(self, ):
        try:
            for file in os.listdir(self._config.resource_path):
                if file.endswith(".xlsx"):
                    os.remove(self._filepath + file)
            return True
        except FileNotFoundError:
            return False

    def delete_specific_xlsx_file(self, filename):
        try:
            os.remove(self._filepath + filename)
            return True
        except FileNotFoundError:
            return False

    def create_worksheet(self, header, data, identifier, filename):
        wb = load_workbook(self._filepath + filename)
        ws = wb.create_sheet(Identifier.get_string_identifier(identifier)[:30])
        if identifier == Identifier.DS_ASSIGNED_TO_CATEGORY or \
                identifier == Identifier.DS_ASSIGNED_TO_COUNTRY_AND_CATEGORY:
            ws = self._write_double_column_data(data, ws)
        else:
            ws = self._write_single_column_data(data, ws)
        wb = self._remove_worksheet("Sheet", wb)
        wb.save(self._filepath + filename)

    def _write_single_column_data(self, data, worksheet):
        for r in dataframe_to_rows(data, index=True, header=True):
            worksheet.append(r)
        worksheet = self._write_extended_formulas_for_single_columns(worksheet)
        worksheet = self._style_worksheet_with_single_columns(worksheet)
        return worksheet

    def _write_extended_formulas_for_single_columns(self, worksheet):
        worksheet = self._fix_rows_after_dataframe_integration(worksheet)
        worksheet = self._write_total_datasets(worksheet)

        last_col = self._get_next_empty_column(worksheet) - 1
        last_row = self._get_next_empty_row(worksheet, last_col) - 1
        worksheet.cell(1, last_col + 1).value = "avg. increase in %"
        worksheet.column_dimensions[self._convert_col_number(last_col)].width = 30
        for row in range(2, last_row + 1):
            fcell = "B" + str(row)
            lcell = self._convert_rowcol(last_col, row)
            worksheet.cell(row, last_col + 1).value = "=IF({0}=0,0,(AVERAGE({0}:{1})-{0})/{0})".format(fcell, lcell)
            worksheet.cell(row, last_col + 1).number_format = "0.00%"
        return worksheet

    def _fix_rows_after_dataframe_integration(self, worksheet, with_deletion=True):
        worksheet.cell(1, 1).value = worksheet.cell(2, 1).value
        if with_deletion:
            worksheet.delete_rows(2)
        else:
            worksheet.cell(2, 1).value = None
        return worksheet

    def _fix_structure_after_dataframe_integration(self, worksheet):
        row = 2
        empty_col = self._get_next_empty_column(worksheet, 1)
        for col in range(2, empty_col):
            if str(worksheet.cell(row - 1, col).value).endswith("1"):
                worksheet.cell(row, col).value = "Assigned DS"
                worksheet.cell(row - 1, col).value = None
            elif str(worksheet.cell(row - 1, col).value).endswith("2"):
                worksheet.cell(row, col).value = "Total DS"
                worksheet.cell(row - 1, col).value = str(worksheet.cell(row - 1, col).value)[:-2]
            elif str(worksheet.cell(row - 1, col).value).endswith("3"):
                worksheet.cell(row, col).value = "in %"
                worksheet.cell(row - 1, col).value = None
        return worksheet

    def _write_total_datasets(self, worksheet, double_column=False):
        last_col = self._get_next_empty_column(worksheet, row=3)
        row = self._get_next_empty_row(worksheet, after_row=3)
        worksheet.cell(row, 1).value = "Total Datasets"
        if not double_column:
            for col in range(2, last_col):
                worksheet.cell(row, col).value = "=SUM({0}:{1})".format(self._convert_rowcol(col, 2),
                                                                        self._convert_rowcol(col, row - 1))
        else:
            for col in range(2, last_col, 3):
                worksheet.cell(row, col).value = "=SUM({0}:{1})".format(self._convert_rowcol(col, 2),
                                                                        self._convert_rowcol(col, row - 1))
                worksheet.cell(row, col + 1).value = "=SUM({0}:{1})".format(self._convert_rowcol(col + 1, 2),
                                                                            self._convert_rowcol(col + 1, row - 1))
                worksheet.cell(row, col + 2).value = "=IF({0}=0,0,{0}/{1})".format(
                    (self._convert_col_number(col) + str(row)),
                    (self._convert_col_number(col + 1) + str(row)))
                col += 2
        return worksheet

    def _write_double_column_data(self, data, worksheet):
        for r in dataframe_to_rows(data, index=True, header=True):
            worksheet.append(r)
        worksheet = self._fix_rows_after_dataframe_integration(worksheet, with_deletion=False)
        worksheet = self._fix_structure_after_dataframe_integration(worksheet)
        worksheet = self._write_total_datasets(worksheet, double_column=True)
        worksheet = self._style_worksheet_with_double_columns(worksheet)
        return worksheet

    def _interpret_value(self, value):
        return 0 if value is None else int(value)

    def _interpret_value_from_double_cols(self, value):
        return value if "NA" in value else int(value)

    def _style_worksheet_with_single_columns(self, worksheet):
        last_col = self._get_next_empty_column(worksheet) - 1
        last_row = self._get_next_empty_row(worksheet) - 1
        worksheet = self._set_line_style(1, 1, last_col, 1, worksheet)
        worksheet = self._set_line_style(1, last_row, last_col, last_row, worksheet)
        worksheet = self._set_column_style(2, last_row - 1, last_col, worksheet)
        worksheet = self._set_date_format(1, 2, last_col, 1, worksheet)
        return worksheet

    def _style_worksheet_with_double_columns(self, worksheet):
        last_col = self._get_next_empty_column(worksheet, 3) - 1
        last_row = self._get_next_empty_row(worksheet, col=3) - 1
        worksheet = self._set_triple_columns_style(2, 1, last_col, last_row, worksheet)
        worksheet = self._set_line_style(1, last_row, last_col, last_row, worksheet)
        worksheet = self._set_double_line_style(1, 1, last_col, 2, worksheet)
        worksheet = self._set_date_format(1, 3, last_col, 3, worksheet)
        return worksheet

    def _create_filename(self, remote_ip):
        date_string = datetime.datetime.now().strftime("%Y-%m-%d_")
        hid = hashlib.md5((date_string + remote_ip).encode())
        return self._fileprefix + date_string + hid.hexdigest()[:6] + self._filesuffix

    def _remove_worksheet(self, title, workbook):
        if title in workbook.sheetnames:
            workbook.remove(workbook[title])
        return workbook

    def _set_double_line_style(self, start_col, start_row, end_col, end_row, worksheet):
        font = Font(name="Calibri", bold=True)
        fill = PatternFill("solid", fgColor="DDDDDD")

        line = Side(border_style="thin", color="000000")
        topborder = Border(top=line)
        bottomoder = Border(bottom=line)

        for col in range(start_col, end_col + 1):
            if worksheet.cell(start_row, col).value is not None:
                worksheet.cell(start_row, col).font = font
            if worksheet.cell(end_row, col).value is not None:
                worksheet.cell(end_row, col).font = font
            worksheet.cell(start_row, col).fill = fill
            worksheet.cell(end_row, col).fill = fill
            worksheet.cell(start_row, col).border = topborder
            worksheet.cell(end_row, col).border = bottomoder

        return worksheet

    def _set_triple_columns_style(self, start_col, start_row, end_col, end_row, worksheet):
        worksheet.column_dimensions["A"].width = 45

        line = Side(border_style="thin", color="000000")
        leftborder = Border(left=line)
        rightborder = Border(right=line)

        for col in range(start_col, end_col + 1, 3):
            worksheet.column_dimensions[self._convert_col_number(col)].width = 11
            worksheet.column_dimensions[self._convert_col_number(col + 1)].width = 11
            worksheet.column_dimensions[self._convert_col_number(col + 2)].width = 11
            for row in range(start_row, end_row + 1):
                worksheet.cell(row, col).border = leftborder
                worksheet.cell(row, col + 2).border = rightborder
                worksheet.cell(row, col + 2).number_format = "0.00%"
                if 3 <= row < end_row:
                    worksheet.cell(row, col + 2).fill = self.__calc_cell_color(worksheet.cell(row, col + 2).value, 0, 100)
        return worksheet

    def __calc_cell_color(self, value, min_value, max_value):
        if value is None:
            return
        d = max_value - min_value
        p = value if value <= 1 else 1
        green = str(hex(int(p * 100 * 2.55)))[2:]
        red = str(hex(255 - int(p * 100 * 2.55)))[2:]
        green = green if len(green) != 1 else "0" + green
        red = red if len(red) != 1 else "0" + red
        colorstring = "{0}{1}00".format(red, green).capitalize()
        return PatternFill("solid", fgColor=colorstring)

    def _set_column_style(self, start_row, end_row, col, worksheet):
        font = Font(color="777777", italic=True)
        for row in range(start_row, end_row + 1):
            worksheet.cell(row, col).font = font
        worksheet.column_dimensions[self._convert_col_number(col)].width = 16
        return worksheet

    def _set_line_style(self, start_col, start_row, end_col, end_row, worksheet):
        font = Font(name="Calibri", bold=True)
        fill = PatternFill("solid", fgColor="DDDDDD")

        line = Side(border_style="thin", color="000000")
        border = Border(top=line, bottom=line)

        for col in range(start_col, end_col + 1):
            worksheet.column_dimensions[self._convert_col_number(col)].width = 11
            for row in range(start_row, end_row + 1):
                worksheet.cell(row, col).font = font
                worksheet.cell(row, col).fill = fill
                worksheet.cell(row, col).border = border
        worksheet.column_dimensions[self._convert_col_number(start_col)].width = 45
        return worksheet

    def _set_date_format(self, row: int, first_col: int, last_col: int, every_col: int, worksheet):
        for col in range(first_col, last_col, every_col):
            date_parts = str(worksheet.cell(row, col).value).split("-")
            worksheet.cell(row, col).value = "{0}.{1}.{2}".format(date_parts[2], date_parts[1], date_parts[0])
        return worksheet

    def _get_next_empty_column(self, worksheet, row=2):
        col = 1
        while worksheet.cell(row, col).value is not None:
            col += 1
        return col

    def _get_next_empty_row(self, worksheet, col=1, after_row=1):
        row = after_row
        while worksheet.cell(row, col).value is not None:
            row += 1
        return row

    def _move_rows(self, moving_value, from_row, col, worksheet):
        last_row_entry = self._get_next_empty_row(worksheet, col) - 1
        worksheet = self._move_cells(from_row, col, last_row_entry + moving_value, col, worksheet)
        return worksheet

    def _move_cells(self, from_row, from_col, to_row, to_col, worksheet):
        for col in range(to_col, from_col - 1, -1):
            last_row_entry = self._get_next_empty_row(worksheet, col) - 1
            for row in range(to_row, from_row - 1, -1):
                worksheet.cell(row, col).value = worksheet.cell(last_row_entry, col).value
                last_row_entry -= 1
            worksheet.cell(last_row_entry + 2, col).value = None
        return worksheet

    def _convert_rowcol(self, col, row):
        return self._convert_col_number(col) + str(row)

    def _convert_col_number(self, col):
        if col <= 26:
            return chr(col + 64)
        if col > 26:
            return "A" + chr(col - 26 + 64)
